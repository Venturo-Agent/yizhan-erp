#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
讀旅遊知識庫 xlsx → 生成 SQL 寫進 knowledge_documents + knowledge_chunks。

設計：
  - 一個地區 sheet → 1 個 document + 11 個 chunks
  - ON CONFLICT UPDATE、可重跑（資料更新時重跑這個 script）
  - metadata 自動從內容推斷標籤（親子 / 銀髮 / 美食 / 季節等）
  - workspace_id 寫死 CORNER（漫途自己用、之後 SaaS 化再參數化）

用法：
  python3 scripts/rag/load-knowledge-from-xlsx.py <xlsx_path> [--region <region>] [--workspace <uuid>]

  --region 只 dump 單一地區 SQL（按 sheet 名稱、譬如 '金澤'）
  --workspace 指定 workspace_id（預設 CORNER）

輸出：SQL 到 stdout、可 redirect 或貼進 MCP execute_sql。
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# 預設 workspace = 漫途 CORNER（handoff 記錄）
DEFAULT_WORKSPACE = 'a89335d4-85f1-492b-83c7-2476ab7c5d81'

# xlsx 「資料欄位」 → DB schema 對應
FIELD_MAP = {
    # 進 documents 的欄位
    '地區名稱':           ('document', 'title'),
    '地區定位標語':       ('document', 'positioning'),
    # 進 chunks 的欄位
    '適合什麼風格的客人': ('chunk', 'audience_fit'),
    '不適合什麼客人':     ('chunk', 'audience_unfit'),
    '核心體驗項目':       ('chunk', 'core_experience'),
    '親子族群注意事項':   ('chunk', 'family_kids'),
    '銀髮族群注意事項':   ('chunk', 'family_senior'),
    '網美打卡族群亮點':   ('chunk', 'instagram'),
    '美食特色':           ('chunk', 'food'),
    '建議天數':           ('chunk', 'duration'),
    '建議搭配地區':       ('chunk', 'pairing'),
    '季節建議與避開時段': ('chunk', 'season'),
    '獨特文化背景':       ('chunk', 'culture'),
}


def parse_region_name(title: str) -> tuple[str, str | None]:
    """
    從「金澤（Kanazawa）— 北陸小京都」抽出 region='金澤', region_en='Kanazawa'。
    """
    m = re.match(r'^([^（(]+)[（(]([^）)]+)[）)]', title.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return title.strip(), None


def detect_country(sheet_name: str) -> str:
    japan_regions = {'金澤', '沖繩', '北海道', '名古屋', '大阪', '東京', '四國',
                     'Kanazawa', 'Okinawa', 'Hokkaido', 'Nagoya', 'Osaka', 'Tokyo', 'Shikoku'}
    thai_regions = {'曼谷', '蘇美島', '清邁',
                    'Bangkok', 'Samui', 'Chiang Mai'}
    for r in japan_regions:
        if r in sheet_name:
            return '日本'
    for r in thai_regions:
        if r in sheet_name:
            return '泰國'
    return '未分類'


# 內容關鍵字 → metadata tags（檢索與篩選用）
AUDIENCE_KEYWORDS = {
    'family_kids':   ['親子', '兒童', '幼童', '小孩', '帶小孩', '寶寶', '主題樂園'],
    'family_senior': ['銀髮', '長輩', '老人', '高齡', '行動不便', '醫療'],
    'couples':       ['情侶', '蜜月', '浪漫', '夜景', '兩人'],
    'friends':       ['閨蜜', '朋友', '姐妹', '青年'],
    'solo':          ['獨旅', '一個人', '單人'],
}

STYLE_KEYWORDS = {
    'leisurely': ['慢活', '悠閒', '靜謐', '療癒', '緩慢'],
    'food':      ['美食', '海鮮丼', '料理', '必吃', '名物', '餐廳'],
    'culture':   ['文化', '傳統', '歷史', '神社', '寺廟', '工藝', '茶道'],
    'nature':    ['自然', '森林', '山', '海', '湖', '溫泉', '花海'],
    'shopping':  ['購物', '百貨', '免稅', 'outlet', 'Outlet'],
    'island':    ['海島', '潛水', '海灘', '碧海'],
    'urban':     ['都會', '繁華', '夜生活', '酒吧'],
}

SEASON_KEYWORDS = {
    'spring': ['春', '櫻花', '4月', '5月'],
    'summer': ['夏', '7月', '8月', '海島'],
    'autumn': ['秋', '楓', '10月', '11月'],
    'winter': ['冬', '雪', '12月', '1月', '2月', '溫泉'],
}


def detect_tags(content: str, category_map: dict) -> list[str]:
    """從文字內容 detect 哪些 tag 命中。"""
    hits = []
    for tag, kws in category_map.items():
        if any(kw in content for kw in kws):
            hits.append(tag)
    return hits


def build_metadata(chunk_type: str, content: str) -> dict:
    """根據 chunk 內容自動推 metadata tags。"""
    md: dict = {}

    # chunk_type 本身就強烈暗示客群
    if chunk_type == 'family_kids':
        md['audience'] = ['family_kids']
    elif chunk_type == 'family_senior':
        md['audience'] = ['family_senior']
    elif chunk_type == 'instagram':
        md['style'] = ['photo']

    # 再從內容補
    aud = detect_tags(content, AUDIENCE_KEYWORDS)
    sty = detect_tags(content, STYLE_KEYWORDS)
    ssn = detect_tags(content, SEASON_KEYWORDS)

    if aud:
        md['audience'] = sorted(set(md.get('audience', []) + aud))
    if sty:
        md['style'] = sorted(set(md.get('style', []) + sty))
    if ssn:
        md['season'] = ssn

    return md


def sql_escape(s: str) -> str:
    """SQL 字串 escape：' → ''"""
    if s is None:
        return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"


def json_to_sql(obj) -> str:
    """jsonb 字面值 → SQL"""
    return sql_escape(json.dumps(obj, ensure_ascii=False)) + '::jsonb'


def load_sheet(ws, sheet_name: str, workspace_id: str) -> str:
    """讀一個地區 sheet、生成該地區 SQL（1 document + N chunks）。"""
    country = detect_country(sheet_name)

    fields: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        k, v = row[0], row[1]
        if not k or not v:
            continue
        fields[str(k).strip()] = str(v).strip()

    title = fields.get('地區名稱', sheet_name)
    region, region_en = parse_region_name(title)
    positioning = fields.get('地區定位標語', '')

    sql_lines: list[str] = []
    sql_lines.append(f'-- ════════ {country} / {region} ════════')

    # 1. INSERT document（ON CONFLICT UPDATE）
    sql_lines.append(f'''WITH upsert_doc AS (
  INSERT INTO public.knowledge_documents (
    workspace_id, country, region, region_en, title, positioning,
    source_file, source_version, metadata
  ) VALUES (
    {sql_escape(workspace_id)}::uuid,
    {sql_escape(country)},
    {sql_escape(region)},
    {sql_escape(region_en)},
    {sql_escape(title)},
    {sql_escape(positioning)},
    {sql_escape("旅遊知識庫_日本泰國_10地區.xlsx")},
    {sql_escape("v1.0")},
    '{{}}'::jsonb
  )
  ON CONFLICT (workspace_id, country, region) DO UPDATE SET
    region_en = EXCLUDED.region_en,
    title = EXCLUDED.title,
    positioning = EXCLUDED.positioning,
    source_file = EXCLUDED.source_file,
    source_version = EXCLUDED.source_version,
    updated_at = now()
  RETURNING id
)''')

    # 2. INSERT chunks（每個 chunk_type 一筆）
    chunk_inserts: list[str] = []
    for field_label, (target, target_key) in FIELD_MAP.items():
        if target != 'chunk':
            continue
        content = fields.get(field_label, '').strip()
        if not content:
            continue
        chunk_type = target_key
        metadata = build_metadata(chunk_type, content)
        chunk_inserts.append(f'''  (
    (SELECT id FROM upsert_doc),
    {sql_escape(workspace_id)}::uuid,
    {sql_escape(chunk_type)},
    {sql_escape(content)},
    {json_to_sql(metadata)}
  )''')

    if chunk_inserts:
        sql_lines.append(f'''INSERT INTO public.knowledge_chunks (
  document_id, workspace_id, chunk_type, content, metadata
) VALUES
{",\n".join(chunk_inserts)}
ON CONFLICT (document_id, chunk_type) DO UPDATE SET
  content = EXCLUDED.content,
  metadata = EXCLUDED.metadata,
  updated_at = now();''')
    else:
        sql_lines.append('-- (no chunks for this region)')

    return '\n'.join(sql_lines)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx_path')
    ap.add_argument('--region', help='Only emit SQL for this region (sheet name fragment)')
    ap.add_argument('--workspace', default=DEFAULT_WORKSPACE)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)

    print('-- ════════════════════════════════════════════════════════════')
    print('-- RAG 知識庫資料載入')
    print(f'-- 來源：{Path(args.xlsx_path).name}')
    print(f'-- workspace_id：{args.workspace}')
    print('-- 可重跑：所有 INSERT 走 ON CONFLICT UPDATE')
    print('-- ════════════════════════════════════════════════════════════')
    print()
    print('BEGIN;')
    print()

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('📋'):
            continue
        if args.region and args.region not in sheet_name:
            continue
        ws = wb[sheet_name]
        print(load_sheet(ws, sheet_name, args.workspace))
        print()

    print('COMMIT;')


if __name__ == '__main__':
    main()
